import io
import json
from http.server import BaseHTTPRequestHandler

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ── 스타일 상수 ───────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="2563EB") if openpyxl else None
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11) if openpyxl else None
_EXCEPTION_FILL = PatternFill("solid", fgColor="FEF3C7") if openpyxl else None
_EXCEPTION_HEADER_FILL = PatternFill("solid", fgColor="D97706") if openpyxl else None

_THIN = Side(style="thin", color="D1D5DB") if openpyxl else None
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN) if openpyxl else None
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True) if openpyxl else None
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True) if openpyxl else None


def _apply_header(cell, value: str, fill=None):
    cell.value = value
    cell.font = _HEADER_FONT
    cell.fill = fill or _HEADER_FILL
    cell.alignment = _CENTER
    cell.border = _BORDER


def _apply_cell(cell, value, align=None):
    cell.value = value
    cell.alignment = align or _CENTER
    cell.border = _BORDER


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 메인 테이블 시트 ──────────────────────────────────────────────────────────

def _write_summary_sheet(ws, summary: list[dict]):
    headers = ["주상병명", "진단코드", "진료기간 (시작)", "진료기간 (종료)", "내원일수", "처치구분", "투약일수"]
    widths = [30, 12, 16, 16, 10, 24, 10]

    ws.row_dimensions[1].height = 30
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _apply_header(ws.cell(row=1, column=col), h)

    _set_col_widths(ws, widths)

    method_label = {"Surgery": "수술", "Treatment": "처치", "Exam": "검사"}

    for row_idx, item in enumerate(summary, 2):
        ws.row_dimensions[row_idx].height = 20
        methods_str = ", ".join(method_label.get(m, m) for m in item.get("methods", []))
        if not methods_str:
            methods_str = "-"

        values = [
            item.get("diagnosis", ""),
            item.get("diagnosisCode", ""),
            item.get("dateRange", {}).get("start", ""),
            item.get("dateRange", {}).get("end", ""),
            item.get("totalVisits", 0),
            methods_str,
            item.get("totalMedicationDays", 0),
        ]
        aligns = [_LEFT, _CENTER, _CENTER, _CENTER, _CENTER, _CENTER, _CENTER]
        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            _apply_cell(ws.cell(row=row_idx, column=col), val, aln)

    # 상단 행 고정
    ws.freeze_panes = "A2"


# ── Exception 시트 ────────────────────────────────────────────────────────────

def _write_exception_sheet(ws, exceptions: list[dict]):
    headers = ["진료일자", "병·의원명", "주상병명", "미매칭 사유"]
    widths = [16, 24, 30, 30]

    ws.row_dimensions[1].height = 30
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _apply_header(ws.cell(row=1, column=col), h, fill=_EXCEPTION_HEADER_FILL)

    _set_col_widths(ws, widths)

    for row_idx, item in enumerate(exceptions, 2):
        ws.row_dimensions[row_idx].height = 20
        ws.cell(row=row_idx, column=1).fill = _EXCEPTION_FILL
        values = [
            item.get("date", ""),
            item.get("hospital", ""),
            item.get("diagnosis", ""),
            item.get("reason", ""),
        ]
        aligns = [_CENTER, _LEFT, _LEFT, _LEFT]
        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            _apply_cell(ws.cell(row=row_idx, column=col), val, aln)
            ws.cell(row=row_idx, column=col).fill = _EXCEPTION_FILL

    ws.freeze_panes = "A2"


# ── 엑셀 생성 ─────────────────────────────────────────────────────────────────

def build_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()

    ws_main = wb.active
    ws_main.title = "보험청구요약"
    _write_summary_sheet(ws_main, data.get("summary", []))

    ws_exc = wb.create_sheet("미매칭항목")
    _write_exception_sheet(ws_exc, data.get("exceptions", []))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Vercel Serverless Handler ─────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        if openpyxl is None:
            self._error(500, "openpyxl not installed")
            return

        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length)

        try:
            data = json.loads(body.decode("utf-8"))
            xlsx_bytes = build_excel(data)

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="insurance_summary.xlsx"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)

        except Exception as e:
            self._error(500, str(e))

    def _error(self, status: int, message: str):
        payload = json.dumps({"error": message}, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def log_message(self, *args):
        pass
